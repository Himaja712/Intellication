from difflib import SequenceMatcher
import requests
import pandas as pd
import io
import numpy as np
import warnings
import re
import tempfile
import urllib.parse
import pytz
import logging
from datetime import datetime, timedelta
import config
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseUpload
from googleapiclient.http import MediaIoBaseDownload

# Global variables from config
DRIVE_CREDENTIALS = config.DRIVE_CREDENTIALS
SCOPES = config.SCOPES
FILE_NAME_TO_DOWNLOAD = config.FILE_NAME_TO_DOWNLOAD
GOOGLE_DRIVE_FOLDER_ID1 = config.GOOGLE_DRIVE_FOLDER_ID1
X_Authorization = config.X_Authorization
status_archived = config.STATUS_ARCHIVED
common_column = config.COMMON_COLUMNS
template_columns = config.template_columns
correct_column_order = config.Merged_column_order
RENAME_MAPPING_EXPENSE = config.RENAME_MAPPING_EXPENSE
EXPENSE_COLUMNS = config.EXPENSE_COLUMNS
RENAME_MAPPING_MILEAGE = config.RENAME_MAPPING_MILEAGE
COLUMN_MAPPING_CONFERENCE = config.COLUMN_MAPPING_CONFERENCE
EXTRA_FIELDS = config.EXTRA_FIELDS
required_cols = config.required_cols
target_fields_expense = config.target_fields_expense
target_fields_mileage = config.target_fields_mileage
target_fields_conference = config.target_fields_conference
target_fields_expense_old = config.target_fields_expense_old
target_fields_mileage_old = config.target_fields_mileage_old
target_fields_conference_old = config.target_fields_conference_old
account_code_1 = config.account_code_1
account_code_2 = config.account_code_2
account_code_3 = config.account_code_3
account_code_1_fields = config.account_code_1_fields
account_code_2_fields = config.account_code_2_fields
account_code_3_fields = config.account_code_3_fields
account_code_1_total = config.account_code_1_total
account_code_2_total = config.account_code_2_total
account_code_3_total = config.account_code_3_total
invoice_field_old = config.invoice_field_old
date_fields_new = config.date_fields_new
invoice_num = config.invoice_number
last = config.last
first = config.first
Emp_id = config.emp_id
escape_header = config.escape_headers
not_found = config.not_found
matched = config.matched
org_id = config.org_id
bank = config.bank
paymt = config.paymt
comment = config.comment
match_col = config.match_col
form = config.form
sheet1 = config.sheet1
sheet2 = config.sheet2
sheet3 = config.sheet3
sheet4 = config.sheet4
sheet5 = config.sheet5
est = config.est
upload_excel_true = config.upload_excel_true

# Handling the Warning related to SettingWithCopyWarning
warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)

# logger called to print the timestamp and message in the terminal
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Initialze the credentials to connect with service account
creds = service_account.Credentials.from_service_account_file(DRIVE_CREDENTIALS, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=creds)

# Get Escape data using folder id and file name
def get_file_id_from_folder(folder_id, file_name):
    query = f"'{folder_id}' in parents and trashed=false"
    try:
        results = drive_service.files().list(
            q=query, 
            fields="files(id, name, mimeType)", 
            includeItemsFromAllDrives=True,  
            supportsAllDrives=True
        ).execute()
        files = results.get('files', [])

        if not files:
            logger.info("No files found in the folder.")
            return None

        for file in files:
            logger.info(f"Found file: {file['name']} ({file['id']})")
            if file['name'] == file_name:  
                return file['id']

        return None 
    
    except Exception as e:
        logger.error(f"Error retrieving file ID: {e}")
        return None

# Extracts the base name from a filename by removing the timestamp if present.
def extract_base_name(filename):
    match = re.search(r'^(.*?)(?: \d{14})?(?:\.\w+)?$', filename)  
    return match.group(1).strip() if match else filename

# get the all the file id from the google drive
def get_all_file_ids_from_folder(folder_id, drive_id=None):
    query = f"'{folder_id}' in parents and trashed=false"
    page_token = None

    try:
        while True:
            response = drive_service.files().list(
                q=query,
                fields="nextPageToken, files(id, name, mimeType)",
                corpora="drive" if drive_id else "user",
                driveId=drive_id if drive_id else None,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
                pageToken=page_token
            ).execute()

            for file in response.get("files", []):
                file_id = file["id"]
                file_name = file["name"]
                mime_type = file["mimeType"]

                try:
                    # Skip folders, only trash files
                    if mime_type != "application/vnd.google-apps.folder":
                        drive_service.files().update(
                            fileId=file_id,
                            body={"trashed": True},
                            supportsAllDrives=True
                        ).execute()
                        logger.info(f"Moved to trash: {file_name}")
                    else:
                        logger.info(f"Skipped folder: {file_name}")

                except Exception as delete_error:
                    logger.warning(f"Failed to delete file {file_name} (ID: {file_id}): {delete_error}")

            page_token = response.get("nextPageToken")
            if not page_token:
                break

    except Exception as e:
        logger.error(f"Error retrieving or deleting files: {e}")

# Download a file from Google Drive and return as a BytesIO stream
def download_google_drive_file(file_id):
    try:
        file_stream = io.BytesIO()
        request = drive_service.files().get_media(fileId=file_id)
        downloader = MediaIoBaseDownload(file_stream, request)
        
        done = False
        while not done:
            _, done = downloader.next_chunk()
        
        file_stream.seek(0)
        return file_stream
    except Exception as e:
        logger.error(f"Fail the download: {e}")
        return None

# Fetch the Informed K12 data from the API for completed status
def fetch_api_data_completed(url, headers):
    all_data = []
    page = 1

    while True:
        try:
            paginated_url = f"{url}&page={page}" if "?" in url else f"{url}?page={page}"
            response = requests.get(paginated_url, headers=headers)
            response.raise_for_status()  # Raise for HTTP errors
            data = response.json()

            if isinstance(data, dict):
                if "data" in data and isinstance(data["data"], list):
                    all_data.extend(data["data"])
                elif "data" in data:
                    all_data.append(data["data"])
                else:
                    all_data.append(data)

                # Handle pagination
                pagination = data.get("meta", {}).get("pagination", {})
                current_page = pagination.get("currentPage", page)
                total_pages = pagination.get("totalPages", page)

                if current_page >= total_pages:
                    break
                page += 1
            elif isinstance(data, list):
                all_data.extend(data)
                break
            else:
                logger.info(f"Unexpected API response format: {data}")
                break
        except Exception as e:
            logger.error(f"Error fetching data from Informed K12 URL (page {page}): {e}")
            break

    return all_data


# Fetch the Informed K12 data from the API for archived status
def fetch_api_data_archived(base_url, campaign_id, headers):
    try:
        EST = pytz.timezone(est)

        # Get current time in EST and 7 weeks ago
        now_est = datetime.now(pytz.utc).astimezone(EST)
        seven_weeks_ago_est = now_est - timedelta(weeks=7)

        # Format as ISO 8601
        completed_at_start = seven_weeks_ago_est.strftime("%Y-%m-%dT%H:%M:%S-07:00")
        completed_at_end = now_est.strftime("%Y-%m-%dT%H:%M:%S-07:00")
        print(completed_at_start)
        print(completed_at_end)

        all_data = []
        page = 1

        while True:
            # Build URL with query string and page number
            endpoint = f"{base_url}{campaign_id}/responses"
            params = {
                "statusGroup": status_archived,
                "completedAtStart": completed_at_start,
                "completedAtEnd": completed_at_end,
                "page": page
            }
            query_string = urllib.parse.urlencode(params)
            full_url = f"{endpoint}?{query_string}"

            # Make request
            response = requests.get(full_url, headers=headers)
            response.raise_for_status()
            data = response.json()

            # Add results from this page
            if "data" in data and isinstance(data["data"], list):
                all_data.extend(data["data"])
            elif "data" in data:
                all_data.append(data["data"])

            # Pagination logic
            pagination = data.get("meta", {}).get("pagination", {})
            current_page = pagination.get("currentPage", page)
            total_pages = pagination.get("totalPages", page)

            if current_page >= total_pages:
                break
            page += 1

        return {"data": all_data}

    except Exception as e:
        logger.error(f"Error fetching archived data: {e}")
        return {"data": []}


# Mapping the number and label to the target fields
def extract_field_mapping(data, target_fields):
    try:
        field_mapping = {}
        if isinstance(data, dict):
            data_array = data.get("data", [])  # Extract list from dictionary
        elif isinstance(data, list):
            data_array = data  # Directly use list if data is already a list
        else:
            logger.info("Unexpected data format")
            return {}

        if not isinstance(data_array, list) or not data_array:
            logger.info("Data array is empty or not a list")
            return {}

        first_item = data_array[0]

        if not isinstance(first_item, dict) or "fields" not in first_item:
            logger.info("First item is not a dictionary or missing 'fields' key")
            return {}

        fields = first_item.get("fields", [])

        # Ensure fields is a list before looping
        if not isinstance(fields, list):
            logger.info("Fields is not a list")
            return {}

        for field in fields:
            field_number = field.get("number")
            field_label = field.get("label")
            if field_number in target_fields:
                field_mapping[field_number] = field_label
        return field_mapping
    except Exception as e:
        logger.error(f"Data array is empty: {e}")


# Load the Escape data into Excel
def load_excel_from_drive(drive_service, file_id):
    try:
        file_stream = io.BytesIO()
        request = drive_service.files().get_media(fileId=file_id)
        downloader = MediaIoBaseDownload(file_stream, request)

        done = False
        while not done:
            _, done = downloader.next_chunk()

        file_stream.seek(0)

        # Ensure we are reading an Excel file correctly
        df_escape = pd.read_excel(file_stream, engine="openpyxl")  
        
        if df_escape is None or df_escape.empty:
            logger.info("Warning: Loaded data is not a DataFrame. Returning an empty DataFrame.")
            return pd.DataFrame() 
        
        logger.info("Successfully loaded Escape data as DataFrame.")
        return df_escape
    
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        return pd.DataFrame() 
    
# Fill missing columns with None/NaN
def ensure_columns(df, required_columns):
    for col in required_columns:
        if col not in df.columns:
            df[col] = None  
    
    return df[required_columns] 

# Rename the expense headers 
def rename_expense_columns(df_expense, field_mapping, EXPENSE_COLUMNS):
    rename_dict = {}

    try:
        # Build rename mapping using field_mapping
        for field_id, new_col in RENAME_MAPPING_EXPENSE.items():
            old_col = field_mapping.get(field_id)
            if old_col and old_col in df_expense.columns:
                rename_dict[old_col] = new_col

        # Apply renaming
        df_expense.rename(columns=rename_dict, inplace=True)

        # Ensure all expected columns are present
        for col in EXPENSE_COLUMNS:
            if col not in df_expense.columns:
                df_expense[col] = None

    except Exception as e:
        logger.error(f"Error in rename_expense_columns: {e}")

    return df_expense


# Rename the mileage headers 
def rename_mileage_columns(df_mileage, field_mapping, common_column):
    rename_dict = {}

    try:
        for field_id, new_col in RENAME_MAPPING_MILEAGE.items():
            old_col = field_mapping.get(field_id)
            if old_col and old_col in df_mileage.columns:
                rename_dict[old_col] = new_col

        df_mileage.rename(columns=rename_dict, inplace=True)

        for col in common_column:
            if col not in df_mileage.columns:
                df_mileage[col] = None 

    except Exception as e:
        logger.error(f"Error in rename_mileage_columns: {e}") 

    return df_mileage
    

# Rename the conference headers 
def rename_conference_columns(df_conference, field_mapping, common_column):
    rename_dict = {}

    try:
        for field_id, new_col in COLUMN_MAPPING_CONFERENCE.items():
            old_col = field_mapping.get(field_id)
            if old_col and old_col in df_conference.columns:
                rename_dict[old_col] = new_col

        df_conference.rename(columns=rename_dict, inplace=True)

        for col in common_column:
            if col not in df_conference.columns:
                df_conference[col] = None  

    except Exception as e:
        logger.error(f"Error in rename_conference_columns: {e}")

    return df_conference

# Rename the Escape headers
def rename_escape_columns(df_escape, escape_header):
    try: 
        rename_dict = {original: new for original, new in zip(df_escape.columns, escape_header)}

        df_escape.rename(columns=rename_dict, inplace=True)
    
    except Exception as e:
        logger.error(f"Error in rename_escape_columns: {e}")
    return ensure_columns(df_escape, escape_header)

# Convert Informed K12 data to Dataframe
def process_api_data(data, field_mapping):
    records = []

    try:
        # Validate input data
        if isinstance(data, dict):
            data_array = data.get("data", [])
        elif isinstance(data, list):
            data_array = data
        else:
            logger.error("Invalid data type passed to process_api_data. Expected dict or list.")
            return pd.DataFrame()

        for i, form in enumerate(data_array):
            form_data = {}
            fields = form.get("fields", [])
            if not isinstance(fields, list):
                logger.warning(f"Invalid 'fields' format in form index {i}. Skipping this entry.")
                continue

            for field in fields:
                try:
                    field_number = field.get("number")
                    if field_number in field_mapping:
                        field_label = field_mapping[field_number]
                        form_data[field_label] = field.get("value", "N/A")
                except Exception as fe:
                    logger.warning(f"Error processing field in form index {i}: {fe}")

            if form_data:
                records.append(form_data)

    except Exception as e:
        logger.error(f"Error in process_api_data: {e}")
        return pd.DataFrame()

    return pd.DataFrame(records)

# Combine the field numbers to get Account codes from conference
def combine_account_codes(df_conference, field_mapping):
    try:
        # Safely map field numbers to field labels
        acc1_cols = [field_mapping.get(num) for num in account_code_1_fields if field_mapping.get(num)]
        acc2_cols = [field_mapping.get(num) for num in account_code_2_fields if field_mapping.get(num)]
        acc3_cols = [field_mapping.get(num) for num in account_code_3_fields if field_mapping.get(num)]

        # Filter columns that actually exist in the DataFrame
        acc1_cols = [col for col in acc1_cols if col in df_conference.columns]
        acc2_cols = [col for col in acc2_cols if col in df_conference.columns]
        acc3_cols = [col for col in acc3_cols if col in df_conference.columns]

        def concat_if_not_empty(row, cols):
            try:
                values = [str(row[col]).strip() for col in cols if pd.notna(row[col]) and str(row[col]).strip()]
                return '-'.join(values) if values else ""
            except Exception as e:
                logger.warning(f"Error concatenating account code columns: {e}")
                return ""

        # Apply concatenation safely
        if acc1_cols:
            df_conference[account_code_1] = df_conference.apply(lambda row: concat_if_not_empty(row, acc1_cols), axis=1)
        if acc2_cols:
            df_conference[account_code_2] = df_conference.apply(lambda row: concat_if_not_empty(row, acc2_cols), axis=1)
        if acc3_cols:
            df_conference[account_code_3] = df_conference.apply(lambda row: concat_if_not_empty(row, acc3_cols), axis=1)

        # Safely drop the original individual columns
        all_cols_to_drop = list(set(acc1_cols + acc2_cols + acc3_cols))
        df_conference.drop(columns=all_cols_to_drop, inplace=True, errors='ignore')

    except Exception as e:
        logger.error(f"Error in combine_account_codes: {e}")

    return df_conference

# def clean_amount(amount):
#     try:
#         if pd.isna(amount):
#             return amount  # Return as-is if NaN

#         amount_str = str(amount).strip()

#         # If the input has no digits at all, treat as 0.00 and log it
#         if not re.search(r"\d", amount_str):
#             return "0.00"

#         # Remove all non-digit and non-dot characters
#         cleaned = re.sub(r"[^\d.]", "", amount_str)

#         # If multiple decimal points, keep only the first one
#         if cleaned.count(".") > 1:
#             parts = cleaned.split(".")
#             cleaned = parts[0] + "." + "".join(parts[1:])

#         # Convert to float and format to 2 decimal places
#         float_val = float(cleaned)
#         return f"{float_val:.2f}"

#     except Exception:
#         return amount 

# Genarate and Fetch Invoice Number
def generate_invoice_number(df, field_mapping, category, is_old_campaign, output_column=invoice_num):
    try:
        if is_old_campaign:
            invoice_field_number = invoice_field_old.get(category)
            invoice_field_label = field_mapping.get(invoice_field_number)

            if invoice_field_label and invoice_field_label in df.columns:
                try:
                    df[output_column] = df[invoice_field_label].astype(str).str.strip()

                    def extract_invoice_or_date(row):
                        try:
                            invoice_number = row.get(output_column, "").strip()
                            if invoice_number:
                                invoice_parts = invoice_number.split()
                                if len(invoice_parts) >= 2:
                                    month = invoice_parts[0].upper()
                                    year = invoice_parts[1]
                                    return f"{month} {year}"
                        except Exception as e:
                            logger.warning(f"Error extracting invoice number from row: {e}")
                        return ""

                    df[output_column] = df.apply(extract_invoice_or_date, axis=1)

                    if invoice_field_label != output_column:
                        df.drop(columns=[invoice_field_label], inplace=True, errors='ignore')
                except Exception as e:
                    logger.error(f"Error processing old campaign invoice number: {e}")
                    df[output_column] = ""
            else:
                df[output_column] = ""
        else:
            date_field_numbers = date_fields_new.get(category, [])
            date_columns = [field_mapping.get(num) for num in date_field_numbers if num in field_mapping]
            date_columns = [col for col in date_columns if col in df.columns]

            def parse_date(value):
                try:
                    return datetime.strptime(str(value).strip(), "%m/%d/%Y")
                except Exception as e:
                    logger.debug(f"Failed to parse date: {value} | Error: {e}")
                    return None

            def compute_invoice(row):
                try:
                    parsed_dates = []
                    for col in date_columns:
                        val = row.get(col)
                        if pd.notna(val) and str(val).strip():
                            date = parse_date(val)
                            if date:
                                parsed_dates.append(date)

                    if not parsed_dates:
                        return ""

                    first_date = min(parsed_dates)
                    last_date = max(parsed_dates)

                    if first_date.year == last_date.year:
                        if first_date.month == last_date.month:
                            return f"{first_date.strftime('%b').upper()} {last_date.year}"
                        else:
                            return f"{first_date.strftime('%b').upper()}-{last_date.strftime('%b').upper()} {last_date.year}"
                    else:
                        if first_date.strftime('%b').upper() == last_date.strftime('%b').upper():
                            return f"{first_date.strftime('%b').upper()} {last_date.year}"
                        else:
                            return f"{first_date.strftime('%b').upper()}-{last_date.strftime('%b').upper()} {last_date.year}"
                except Exception as e:
                    logger.warning(f"Error computing invoice for row: {e}")
                    return ""

            try:
                df[output_column] = df.apply(compute_invoice, axis=1)
                df.drop(columns=date_columns, inplace=True, errors='ignore')
            except Exception as e:
                logger.error(f"Error processing new campaign invoice number: {e}")
                df[output_column] = ""

    except Exception as e:
        logger.critical(f"Critical error in generate_invoice_number for category '{category}': {e}")
        df[output_column] = ""

    return df


# Remove the f/0 from the employee id
def clean_employee_id(emp_id):
    try:
        if pd.isna(emp_id) or str(emp_id).strip() == "":
            return ""  # Keep it blank if NA or empty

        # Remove all non-digit characters
        emp_id_cleaned = re.sub(r'\D', '', str(emp_id))

        if emp_id_cleaned.isdigit():
            return int(emp_id_cleaned)  # Return as integer
        else:
            return ""
    except Exception as e:
        logger.warning(f"Error cleaning employee ID '{emp_id}': {e}")
        return ""

# Delete the row if no data present
def clean_and_drop_empty_rows(df):
    def is_blank(x):
        return pd.isna(x) or str(x).strip() == ""

    # Apply blank replacement
    df = df.apply(lambda col: col.map(lambda x: "" if is_blank(x) else x))
    df = df[~df.apply(lambda row: all(is_blank(x) for x in row), axis=1)]

    return df

# Create a dataframe using Emp Id from Informed K12 data and Escape data
def prepare_dataframes(df_expense, df_escape):
    try:
        df_expense = clean_and_drop_empty_rows(df_expense)
    except Exception as e:
        logger.error(f"Error cleaning and dropping rows from df_expense: {e}")
        return df_expense, df_escape  # Early return if essential preprocessing fails

    try:
        # Handle potentially invalid employee IDs
        invalid_emp_rows = df_expense[
            df_expense.get(Emp_id, pd.Series(dtype=object)).apply(lambda x: isinstance(x, str) and not x.isdigit())
        ].copy()
        if not invalid_emp_rows.empty:
            invalid_emp_rows[match_col] = not_found

        df_expense[Emp_id] = df_expense.get(Emp_id, pd.Series(dtype=object)).apply(clean_employee_id)
        df_escape[Emp_id] = df_escape.get(Emp_id, pd.Series(dtype=object)).apply(clean_employee_id)

        df_expense[Emp_id] = df_expense[Emp_id].replace("", pd.NA)
    except Exception as e:
        logger.warning(f"Error cleaning employee IDs: {e}")

    try:
        df_expense[Emp_id] = df_expense[Emp_id].astype("Int64")
    except Exception as e:
        logger.warning(f"Error converting df_expense[Emp_id] to Int64: {e}")

    try:
        df_escape[Emp_id] = df_escape[Emp_id].astype("Int64")
    except Exception as e:
        logger.warning(f"Error converting df_escape[Emp_id] to Int64: {e}")

    return df_expense, df_escape

# Comparison based on Emp Id and highlight is based on Email.
def check_match(row):
    if pd.isna(row.get(Emp_id)) or pd.isna(row.get("Email_Escape")):
        return not_found
    return matched

# Merge the Escape, Mileage and Conference with Expense data, and compare them
def merge_and_compare_common(df_source, df_escape, form_type, include_account_code_3=True, invalid_emp_rows=None):
    try:
        df_source[Emp_id] = df_source[Emp_id].astype(str)
        df_escape[Emp_id] = df_escape[Emp_id].astype(str)
    except Exception as e:
        logger.error(f"Error converting Employee IDs to string: {e}")
        return pd.DataFrame()

    try:
        comparison_df = df_source.merge(
            df_escape,
            on=Emp_id,
            how="left",
            indicator=True
        ).copy()
    except Exception as e:
        logger.error(f"Error merging dataframes: {e}")
        return pd.DataFrame()

    try:
        # Fill missing names
        comparison_df["Last_y"] = comparison_df.get("Last_y", np.nan).fillna(comparison_df.get("Last_x"))
        comparison_df["First_y"] = comparison_df.get("First_y", np.nan).fillna(comparison_df.get("First_x"))

        # Match logic
        comparison_df[match_col] = comparison_df.apply(check_match, axis=1)

        # Fix columns
        comparison_df.rename(columns={"Last_y": last, "First_y": first}, inplace=True)
    except Exception as e:
        logger.warning(f"Error applying name fix and match logic: {e}")

    # Handle optional fields
    extra_fields = [col for col in EXTRA_FIELDS if col in df_source.columns]

    if include_account_code_3:
        for col in [account_code_3, account_code_3_total]:
            if col not in comparison_df.columns:
                comparison_df[col] = None

    comparison_df[form] = form_type

    final_columns = required_cols + extra_fields
    if include_account_code_3:
        final_columns += [account_code_3, account_code_3_total]
    final_columns += [form]
    final_columns = [col for col in final_columns if col in comparison_df.columns]

    # Handle invalid employee rows if provided
    try:
        if invalid_emp_rows is not None:
            for col in final_columns:
                if col not in invalid_emp_rows.columns:
                    invalid_emp_rows[col] = None

            invalid_emp_rows[form] = form_type
            matched_ids = comparison_df[Emp_id].astype(str).unique()
            invalid_emp_rows = invalid_emp_rows[~invalid_emp_rows[Emp_id].astype(str).isin(matched_ids)]
            invalid_emp_rows_final = invalid_emp_rows[final_columns]
        else:
            invalid_emp_rows_final = pd.DataFrame(columns=final_columns)
    except Exception as e:
        logger.warning(f"Error processing invalid employee rows: {e}")
        invalid_emp_rows_final = pd.DataFrame(columns=final_columns)

    # Final assembly
    try:
        comparison_df_final = comparison_df[final_columns]
        dfs_to_concat = [comparison_df_final]

        if not invalid_emp_rows_final.empty and not invalid_emp_rows_final.isna().all().all():
            dfs_to_concat.append(invalid_emp_rows_final)

        final_df = pd.concat(dfs_to_concat, ignore_index=True)
    except Exception as e:
        logger.error(f"Error concatenating comparison results: {e}")
        return pd.DataFrame()

    try:
        # Clean up and deduplicate
        final_df[first] = final_df[first].replace(["Nan", "nan"], np.nan)
        final_df[last] = final_df[last].replace(["Nan", "nan"], np.nan)

        final_df = final_df[~(
            final_df[Emp_id].isna() &
            final_df[first].isna() &
            final_df[last].isna()
        )]

        final_df[first] = final_df[first].astype(str).str.strip().str.lower()
        final_df[last] = final_df[last].astype(str).str.strip().str.lower()

        df_with_id = final_df[final_df[Emp_id].notna()]
        df_without_id = final_df[final_df[Emp_id].isna()].drop_duplicates(subset=[first, last, form])

        final_df = pd.concat([df_with_id, df_without_id], ignore_index=True)

        final_df[first] = final_df[first].str.title()
        final_df[last] = final_df[last].str.title()

        comparison_df = final_df
        return comparison_df
    except Exception as e:
        logger.error(f"Final cleanup or formatting error: {e}")
        return pd.DataFrame()


# Merge and comapre Expense and Escape
def merge_and_compare_data_expense(df_expense, df_escape, invalid_emp_rows=None):
    return merge_and_compare_common(df_expense, df_escape, form_type="Expense", include_account_code_3=True, invalid_emp_rows=invalid_emp_rows)

# Merge and comapre Mileage and Escape
def merge_and_compare_data_mileage(df_mileage, df_escape, invalid_emp_rows=None):
    return merge_and_compare_common(df_mileage, df_escape, form_type="Mileage", include_account_code_3=True, invalid_emp_rows=invalid_emp_rows)

# Merge and comapre Conference and Escape
def merge_and_compare_data_conference(df_conference, df_escape, invalid_emp_rows=None):
    return merge_and_compare_common(df_conference, df_escape, form_type="Conference", include_account_code_3=False, invalid_emp_rows=invalid_emp_rows)

# Merge all the comparison_df to get final_comparison_df
def merge_and_compare_data_combined(df_expense, df_escape, df_mileage, df_conference, invalid_emp_rows=None):
    try:
        comparison_df = merge_and_compare_data_expense(df_expense, df_escape, invalid_emp_rows=invalid_emp_rows)
    except Exception as e:
        logger.error(f"Error in merge_and_compare_data_expense: {e}")
        comparison_df = pd.DataFrame(columns=correct_column_order)

    try:
        comparison_df1 = merge_and_compare_data_mileage(df_mileage, df_escape, invalid_emp_rows=invalid_emp_rows)
    except Exception as e:
        logger.error(f"Error in merge_and_compare_data_mileage: {e}")
        comparison_df1 = pd.DataFrame(columns=correct_column_order)

    try:
        comparison_df2 = merge_and_compare_data_conference(df_conference, df_escape, invalid_emp_rows=invalid_emp_rows)
    except Exception as e:
        logger.error(f"Error in merge_and_compare_data_conference: {e}")
        comparison_df2 = pd.DataFrame(columns=correct_column_order)

    try:
        all_columns = set(correct_column_order)

        for df in [comparison_df, comparison_df1, comparison_df2]:
            missing_cols = all_columns - set(df.columns)
            for col in missing_cols:
                df[col] = None

        comparison_df = comparison_df[correct_column_order]
        comparison_df1 = comparison_df1[correct_column_order]
        comparison_df2 = comparison_df2[correct_column_order]
    except Exception as e:
        logger.error(f"Error aligning columns: {e}")
        return pd.DataFrame(columns=correct_column_order)

    try:
        for df in [comparison_df, comparison_df1, comparison_df2]:
            df.loc[:, match_col] = df.apply(lambda row: check_match(row), axis=1)
    except Exception as e:
        logger.warning(f"Error applying match logic: {e}")

    try:
        final_comparison_df = pd.concat([comparison_df, comparison_df1, comparison_df2], ignore_index=True)
    except Exception as e:
        logger.error(f"Error concatenating comparison DataFrames: {e}")
        final_comparison_df = pd.DataFrame(columns=correct_column_order)

    return final_comparison_df

# Remove the 0 and / from the Invoice date
def format_invoice_date(date_value):
    try:
        # Early return for clearly invalid types
        if pd.isna(date_value) or not isinstance(date_value, (str, int, float, pd.Timestamp)):
            return ""

        date_str = str(date_value).strip()
        if not date_str:
            return ""

        # Convert to datetime
        date_obj = pd.to_datetime(date_str, errors="coerce")

        if pd.isna(date_obj):
            logger.warning(f"Unable to parse date: {date_value}")
            return ""

        # Format: MDDYYYY (remove leading zero from month, no slashes)
        return f"{date_obj.month}{date_obj.day:02d}{date_obj.year}"

    except Exception as e:
        logger.error(f"Unexpected error formatting date '{date_value}': {e}")
        return ""


# Fetch the "JUL 2024"
def mon_year_invoice_date(date):
    if not date or pd.isna(date):
        return ""

    try:
        date_str = str(date).strip()

        if not date_str:
            return ""

        # Handle 7/8-digit compact formats like 012024 or 01012024
        if re.match(r"^\d{7,8}$", date_str):
            if len(date_str) == 7:
                date_str = f"0{date_str[:1]}/{date_str[1:3]}/{date_str[3:]}"
            elif len(date_str) == 8:
                date_str = f"{date_str[:2]}/{date_str[2:4]}/{date_str[4:]}"

        # Try parsing using dd/mm/yyyy first
        date_obj = pd.to_datetime(date_str, format="%d/%m/%Y", errors="coerce")

        if pd.isna(date_obj):
            # Fallback: attempt general parsing (e.g., mm/dd/yyyy, ISO, etc.)
            date_obj = pd.to_datetime(date_str, format="%m/%d/%Y", errors="coerce")

        if pd.notna(date_obj):
            return date_obj.strftime("%b %Y").upper()

    except Exception as e:
        logger.error(f"Error formatting date {date}: {e}")

    return ""

# Remove any characters that are not digits or dashes and cleaned account number matches the expected format
def format_account_number(account):
    if account is None:
        return ""

    try:
        account = str(account).strip()

        # Accept if it's a non-numeric string like "BILL SUNOL GLEN"
        if not any(char.isdigit() for char in account):
            return account

        # Replace periods with dashes
        account = account.replace(".", "-")

        # Remove any characters that are not digits or dashes
        account = re.sub(r"[^0-9\-]", "", account)

        # Check if the cleaned account number matches the expected format
        expected_format = r"^\d{3}-\d{4}-\d-\d{4}-\d{4}-\d{4}-\d{3}-\d{4}-\d{4}$"
        if re.fullmatch(expected_format, account):
            return account
        else:
            return ""

    except Exception as e:
        logger.error(f"Error formatting account number '{account}': {e}")
        return ""
    


# Empty, string or invalid account code will be highlighted red
def needs_red_highlight(account):
    if not account or str(account).strip() == "":
        return True
    if not any(char.isdigit() for char in str(account)):
        return True
    expected_format = r"^\d{3}-\d{4}-\d-\d{4}-\d{4}-\d{4}-\d{3}-\d{4}-\d{4}$"
    return not re.match(expected_format, str(account))

# If amount is not empty then the account code will be highlighted. 
def should_highlight_account(row):
    # Check account_code_1
    total_1_raw = str(row.get(account_code_1_total, "")).replace(",", "")
    total_1 = pd.to_numeric(total_1_raw, errors="coerce")
    if not pd.isna(total_1):
        acc_1 = row.get(account_code_1, "")
        if needs_red_highlight(acc_1):
            return True

    # Check account_code_2
    total_2_raw = str(row.get(account_code_2_total, "")).replace(",", "")
    total_2 = pd.to_numeric(total_2_raw, errors="coerce")
    if not pd.isna(total_2):
        acc_2 = row.get(account_code_2, "")
        if needs_red_highlight(acc_2):
            return True

    # Check account_code_3
    total_3_raw = str(row.get(account_code_3_total, "")).replace(",", "")
    total_3 = pd.to_numeric(total_3_raw, errors="coerce")
    if not pd.isna(total_3):
        acc_3 = row.get(account_code_3, "")
        if needs_red_highlight(acc_3):
            return True

    return False

# Create the csv file using the merged sheet
def create_matched_data_sheet(comparison_df):
    try:
        if comparison_df.empty:
            return pd.DataFrame(columns=template_columns)

        # Highlight accounts (catch errors per row)
        try:
            comparison_df["Highlight_Account"] = comparison_df.apply(should_highlight_account, axis=1)
        except Exception as e:
            logger.error(f"Error applying should_highlight_account: {e}")
            comparison_df["Highlight_Account"] = False

        # Filter matched or not_found rows
        matched_df = comparison_df[comparison_df[match_col].isin([matched, not_found])].copy()

        if form in matched_df.columns:
            matched_df.drop(columns=[form], inplace=True)

        # Prepare the column mapping for new columns
        column_mapping = {
            "Payee Name": matched_df[first].fillna('') + " " + matched_df[last].fillna(''),
        }

        if Emp_id in comparison_df.columns:
            matched_df["EmpId"] = matched_df[Emp_id]
        else:
            logger.warning("Employee ID column is missing from comparison_df.")
            matched_df["EmpId"] = None

        if "Invoice Date" in comparison_df.columns:
            try:
                matched_df["Invoice Date"] = matched_df["Invoice Date"].apply(format_invoice_date)
            except Exception as e:
                logger.error(f"Error formatting 'Invoice Date': {e}")
                matched_df["Invoice Date"] = matched_df["Invoice Date"]
        else:
            logger.warning("Invoice Date column is missing from comparison_df.")
            matched_df["Invoice Date"] = None

        # Ensure match_col is in template_columns
        if match_col not in template_columns:
            template_columns.append(match_col)

        expanded_rows = []

        for idx, row in matched_df.iterrows():
            try:
                row_copy = row.copy()

                invoice_number = str(row_copy.get("Invoice #", "")).strip()
                invoice_date = str(row_copy.get("Invoice Date", "")).strip()

                if invoice_number:
                    invoice_base = invoice_number
                elif invoice_date:
                    invoice_base = mon_year_invoice_date(invoice_date)
                else:
                    invoice_base = ""

                amount_1 = str(row_copy.get(account_code_1_total, "0")).replace(",", "")
                amount_2 = str(row_copy.get(account_code_2_total, "0")).replace(",", "")
                amount_3 = str(row_copy.get(account_code_3_total, "0")).replace(",", "")
                total_reimbursement = str(row_copy.get("Total Reimbursement", "0")).replace(",", "")

                # amount_1 = clean_amount(row_copy.get(account_code_1_total, "0"))
                # amount_2 = clean_amount(row_copy.get(account_code_2_total, "0"))
                # amount_3 = clean_amount(row_copy.get(account_code_3_total, "0"))
                # total_reimbursement = clean_amount(row_copy.get("Total Reimbursement", "0"))


                amount_1 = pd.to_numeric(amount_1, errors="coerce") or 0.0
                amount_2 = pd.to_numeric(amount_2, errors="coerce") or 0.0
                amount_3 = pd.to_numeric(amount_3, errors="coerce") or 0.0
                total_reimbursement = pd.to_numeric(total_reimbursement, errors="coerce") or 0.0

                account_amounts = {}
                original_accounts = [
                    (row_copy.get(account_code_1, ""), amount_1),
                    (row_copy.get(account_code_2, ""), amount_2),
                    (row_copy.get(account_code_3, ""), amount_3)
                ]

                has_account = False

                for raw_acc, amt in original_accounts:
                    formatted_acc = ""
                    try:
                        formatted_acc = format_account_number(raw_acc)
                    except Exception as e:
                        logger.error(f"Error formatting account number '{raw_acc}': {e}")

                    final_acc = formatted_acc if formatted_acc else (raw_acc if raw_acc and str(raw_acc).strip() else "")

                    if str(final_acc).strip() and pd.notna(amt):
                        has_account = True
                        if final_acc in account_amounts:
                            account_amounts[final_acc] += amt
                        else:
                            account_amounts[final_acc] = amt

                # Logic for only one account with total reimbursement > 0
                if has_account and len(account_amounts) == 1 and total_reimbursement > 0:
                    only_account = next(iter(account_amounts))
                    account_amounts[only_account] = total_reimbursement

                # Add rows based on accounts and amounts
                if has_account:
                    for account, amount in account_amounts.items():
                        row_copy_new = row.copy()
                        row_copy_new["Account"] = account
                        row_copy_new["Amount"] = amount
                        row_copy_new["Invoice #"] = f"{invoice_base} {int(amount * 100)}" if invoice_base else ""
                        row_copy_new[match_col] = row.get(match_col, "")
                        try:
                            row_copy_new["Highlight_Account"] = needs_red_highlight(account)
                        except Exception as e:
                            logger.error(f"Error determining highlight for account '{account}': {e}")
                            row_copy_new["Highlight_Account"] = False
                        expanded_rows.append(row_copy_new)

                elif total_reimbursement > 0:
                    row_copy_new = row.copy()
                    # raw_acc_fallback = row_copy.get(account_code_1, "") or row_copy.get(account_code_2, "") or row_copy.get(account_code_3, "")
                    # row_copy_new["Account"] = raw_acc_fallback if raw_acc_fallback and str(raw_acc_fallback).strip() else " "
                    row_copy_new["Account"] = " "
                    row_copy_new["Amount"] = total_reimbursement
                    row_copy_new["Invoice #"] = f"{invoice_base} {int(total_reimbursement * 100)}" if invoice_base else ""
                    row_copy_new[match_col] = row.get(match_col, "")
                    row_copy_new["Highlight_Account"] = True
                    expanded_rows.append(row_copy_new)

                else:
                    row_copy_new = row.copy()
                    raw_acc_fallback = row_copy.get(account_code_1, "") or row_copy.get(account_code_2, "") or row_copy.get(account_code_3, "")
                    row_copy_new["Account"] = raw_acc_fallback if raw_acc_fallback and str(raw_acc_fallback).strip() else " "
                    row_copy_new["Amount"] = 0
                    row_copy_new["Invoice #"] = f"{invoice_base} 0" if invoice_base else ""
                    row_copy_new[match_col] = row.get(match_col, "")
                    try:
                        row_copy_new["Highlight_Account"] = needs_red_highlight(raw_acc_fallback)
                    except Exception as e:
                        logger.error(f"Error determining highlight for fallback account '{raw_acc_fallback}': {e}")
                        row_copy_new["Highlight_Account"] = False
                    expanded_rows.append(row_copy_new)

            except Exception as e:
                logger.error(f"Error processing row index {idx}: {e}")

        matched_df = pd.DataFrame(expanded_rows)

        # Rename and create columns based on mapping
        for new_col, old_col in column_mapping.items():
            if isinstance(old_col, pd.Series):
                try:
                    matched_df[new_col] = old_col
                except Exception as e:
                    logger.error(f"Error assigning series to column '{new_col}': {e}")
                    matched_df[new_col] = None
            elif old_col in matched_df.columns:
                matched_df[new_col] = matched_df[old_col]
            else:
                matched_df[new_col] = None

        # Add default columns and values safely
        try:
            matched_df["Org ID"] = org_id
        except Exception:
            matched_df["Org ID"] = None
            logger.warning("org_id variable not found; 'Org ID' column set to None.")

        try:
            matched_df["Amount"] = matched_df["Amount"].astype(float)
        except Exception as e:
            logger.error(f"Error converting 'Amount' to float: {e}")

        try:
            matched_df["Bank"] = bank
        except Exception:
            matched_df["Bank"] = None
            logger.warning("bank variable not found; 'Bank' column set to None.")

        try:
            matched_df["PymtType"] = paymt
        except Exception:
            matched_df["PymtType"] = None
            logger.warning("paymt variable not found; 'PymtType' column set to None.")

        try:
            matched_df["Comment"] = comment
        except Exception:
            matched_df["Comment"] = None
            logger.warning("comment variable not found; 'Comment' column set to None.")

        # Ensure all template columns exist
        for col in template_columns:
            if col not in matched_df.columns:
                matched_df.loc[:, col] = None

        # Reorder columns based on template
        matched_df = matched_df[template_columns]

        return matched_df

    except Exception as e:
        logger.error(f"Fatal error in create_matched_data_sheet: {e}")
        # Return empty DataFrame with template columns on fatal failure
        return pd.DataFrame(columns=template_columns)
    

# Apply highlighting to the sheet based on Match Status.
def apply_highlighting(excel_file_path, include_legend=True, highlight=True):
    wb = load_workbook(excel_file_path)
    ws = wb[sheet5]

    # Define highlight colors
    red_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    match_status_col = None
    account_code_cols = []
    payee_name_col = None

    # Identify required column indices
    for col in range(1, ws.max_column + 1):
        col_value = ws.cell(row=1, column=col).value
        if col_value == match_col:
            match_status_col = col
        elif isinstance(col_value, str) and col_value.startswith("Account") and "Total" not in col_value:
            account_code_cols.append(col)
        elif col_value == "Payee Name":
            payee_name_col = col

    if match_status_col is None:
        return {"error": "Match Status column not found!"}

    # Highlight rows with missing Payee Name
    if payee_name_col:
        for row in range(2, ws.max_row + 1):
            payee_name = ws.cell(row=row, column=payee_name_col).value
            if not str(payee_name).strip():
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill

    # Highlight rows based on Match Status
    for row in range(2, ws.max_row + 1):
        match_status = ws.cell(row=row, column=match_status_col).value
        if match_status == not_found:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill

    # Highlight invalid Account Codes
    if highlight:
        for row in range(2, ws.max_row + 1):
            for col in account_code_cols:
                account_value = ws.cell(row=row, column=col).value
                account_str = str(account_value).strip() if account_value is not None else ""

                if not format_account_number(account_value) or needs_red_highlight(account_str):
                    ws.cell(row=row, column=col).fill = red_fill
    else:
        # Use column names to find account/total pairings
        account_code_cols = [
            correct_column_order.index("Account Code 1") + 1,
            correct_column_order.index("Account Code 2") + 1,
            correct_column_order.index("Account Code 3") + 1,
        ]

        account_col_to_total_col = {
            account_code_cols[0]: account_code_1_total,
            account_code_cols[1]: account_code_2_total,
            account_code_cols[2]: account_code_3_total,
        }

        for row in range(2, ws.max_row + 1):
            for col in account_code_cols:
                account_value = ws.cell(row=row, column=col).value
                total_col_name = account_col_to_total_col.get(col)
                total_col_index = correct_column_order.index(total_col_name) + 1
                total_value = ws.cell(row=row, column=total_col_index).value
                total_numeric = pd.to_numeric(str(total_value).replace(",", ""), errors="coerce") if total_value else None

                if pd.isna(total_numeric):
                    continue

                account_str = str(account_value).strip() if account_value is not None else ""
                if needs_red_highlight(account_str):
                    ws.cell(row=row, column=col).fill = red_fill

    # Delete "Highlight_Account" column if present
    highlight_col_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Highlight_Account":
            highlight_col_index = col
            break
    if highlight_col_index:
        ws.delete_cols(highlight_col_index)

    # Delete Match Status column
    ws.delete_cols(match_status_col)

    # Apply green header highlight
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).fill = green_fill

    # Add legend if needed
    if include_legend:
        last_column = ws.max_column
        label_col = last_column + 4
        ws.cell(row=1, column=label_col, value="").fill = red_fill
        ws.cell(row=1, column=label_col + 1, value=not_found + " (Employee ID Missing or Incorrect Account Code Format)")
        ws.cell(row=2, column=label_col + 1, value=matched)

    wb.save(excel_file_path)
    wb.close()

# Save processed data to an Excel file. 
def save_to_excel(df_expense, df_mileage, df_conference, df_escape, final_comparison_df):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_excel_path = temp_file.name

    with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as writer:
        df_expense.to_excel(writer, sheet_name=sheet1, index=False)
        df_mileage.to_excel(writer, sheet_name=sheet2, index=False)
        df_conference.to_excel(writer, sheet_name=sheet3, index=False)
        df_escape.to_excel(writer, sheet_name=sheet4, index=False)
        final_comparison_df.to_excel(writer, sheet_name=sheet5, index=False)
    apply_highlighting(temp_excel_path, include_legend=True, highlight=False) # Apply the highlight
    return temp_excel_path

# Save final data to an Excel file. 
def save_to_excel_final(matched_df):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_csv_file:
            temp_csv_path = temp_csv_file.name

    with pd.ExcelWriter(temp_csv_path, engine='openpyxl') as writer:
        matched_df.to_excel(writer, sheet_name=sheet5, index=False)
    apply_highlighting(temp_csv_path, include_legend=False, highlight=True)
    return temp_csv_path
    
# Find the next Sunday at 23:59:59
def get_current_timestamp():
    now = datetime.now(pytz.timezone(est))  # Get current EST time
    return now.strftime("%Y%m%d%H%M%S")

def get_date_range_filename():
    # Get current EST time
    now = datetime.now(pytz.timezone(est))
    
    # Get the starting date (7 weeks ago)
    start_date = now - timedelta(weeks=7)

    # Format the dates as YYYYMMDD
    start_date_str = start_date.strftime("%Y%m%d")
    end_date_str = now.strftime("%Y%m%d")

    # Return formatted filename
    return start_date_str, end_date_str

# Process to fetch data, transform and upload back to google drive.
def process_and_upload_files(old_expense_data, new_expense_data, old_mileage_data, new_mileage_data, old_conference_data, new_conference_data, drive_service, GOOGLE_DRIVE_FOLDER_ID1, GOOGLE_DRIVE_FOLDER_ID2, FILE_NAME_TO_DOWNLOAD, X_Authorization, new_excel_name, new_csv_name, upload_excel_true):
    try:
        logger.info("Creating new file..")
        file_id = get_file_id_from_folder(GOOGLE_DRIVE_FOLDER_ID1, FILE_NAME_TO_DOWNLOAD)

        if file_id:
            df_escape = load_excel_from_drive(drive_service, file_id)
            logger.info("Fetched the Escape data")
            if df_escape.empty:
                logger.info(f"Warning: File {FILE_NAME_TO_DOWNLOAD} is empty. Initializing an empty DataFrame.")
                df_escape = pd.DataFrame()  # Ensure df_escape is empty if the file has no data
        else:
            logger.info(f"File '{FILE_NAME_TO_DOWNLOAD}' not found in folder {GOOGLE_DRIVE_FOLDER_ID1}. Creating an empty Escape sheet.")
            df_escape = pd.DataFrame()

        logger.info("Fetched the API data")

        new_expense_mapping = extract_field_mapping(new_expense_data, target_fields_expense)
        old_expense_mapping = extract_field_mapping(old_expense_data, target_fields_expense_old)

        df_expense_new = process_api_data(new_expense_data, new_expense_mapping)
        df_expense_new = generate_invoice_number(df_expense_new, new_expense_mapping, 'expense', is_old_campaign=False)
        df_expense_new = rename_expense_columns(df_expense_new, new_expense_mapping, EXPENSE_COLUMNS)

        df_expense_old = process_api_data(old_expense_data, old_expense_mapping)
        df_expense_old = generate_invoice_number(df_expense_old, old_expense_mapping, 'expense', is_old_campaign=True)
        df_expense_old = rename_expense_columns(df_expense_old, old_expense_mapping, EXPENSE_COLUMNS)

        df_expense = pd.concat([df_expense_old, df_expense_new], ignore_index=True)

        new_mileage_mapping = extract_field_mapping(new_mileage_data, target_fields_mileage)
        old_mileage_mapping = extract_field_mapping(old_mileage_data, target_fields_mileage_old)

        df_mileage_new = process_api_data(new_mileage_data, new_mileage_mapping)
        df_mileage_new = generate_invoice_number(df_mileage_new, new_mileage_mapping, 'mileage', is_old_campaign=False)
        df_mileage_new = rename_mileage_columns(df_mileage_new, new_mileage_mapping, common_column)

        df_mileage_old = process_api_data(old_mileage_data, old_mileage_mapping)
        df_mileage_old = generate_invoice_number(df_mileage_old, old_mileage_mapping, 'mileage', is_old_campaign=True)
        df_mileage_old = rename_mileage_columns(df_mileage_old, old_mileage_mapping, common_column)

        df_mileage = pd.concat([df_mileage_old, df_mileage_new], ignore_index=True)

        new_conference_mapping = extract_field_mapping(new_conference_data, target_fields_conference)
        old_conference_mapping = extract_field_mapping(old_conference_data, target_fields_conference_old)

        df_conference_new = process_api_data(new_conference_data, new_conference_mapping)
        df_conference_new = generate_invoice_number(df_conference_new, new_conference_mapping, 'conference', is_old_campaign=False)
        df_conference_new = combine_account_codes(df_conference_new, new_conference_mapping)
        df_conference_new = rename_conference_columns(df_conference_new, new_conference_mapping, common_column)

        df_conference_old = process_api_data(old_conference_data, old_conference_mapping)
        df_conference_old = generate_invoice_number(df_conference_old, old_conference_mapping, 'conference', is_old_campaign=True)
        df_conference_old = combine_account_codes(df_conference_old, old_conference_mapping)
        df_conference_old = rename_conference_columns(df_conference_old, old_conference_mapping, common_column)

        df_conference= pd.concat([df_conference_old, df_conference_new], ignore_index=True)

        if df_escape.empty:
            logger.info("Escape data is empty. Initializing with headers only.")
            df_escape = pd.DataFrame(columns=escape_header)  # Create an empty DataFrame with these headers
            final_comparison_df = pd.DataFrame(columns=correct_column_order)  # Empty Sheet3
            matched_df = create_matched_data_sheet(final_comparison_df) 

        else:
            df_escape = rename_escape_columns(df_escape, escape_header)
            
            if df_expense.empty:
                logger.info("Expense data is empty. Initializing with headers only.")
                df_expense = pd.DataFrame(columns=EXPENSE_COLUMNS)
            else:
                df_expense, df_escape = prepare_dataframes(df_expense, df_escape)

            if df_mileage.empty:
                logger.info("Mileage data is empty. Initializing with headers only.")
                df_mileage = pd.DataFrame(columns=common_column)
            else:
                df_mileage, df_escape = prepare_dataframes(df_mileage, df_escape)

            if df_conference.empty:
                logger.info("Conference data is empty. Initializing with headers only.")
                df_conference = pd.DataFrame(columns=common_column)
            else:
                df_conference, df_escape = prepare_dataframes(df_conference, df_escape)

            final_comparison_df = merge_and_compare_data_combined(df_expense, df_escape, df_mileage, df_conference)
            matched_df = create_matched_data_sheet(final_comparison_df)

        temp_excel_path = save_to_excel(df_expense, df_mileage, df_conference, df_escape, final_comparison_df)
        temp_csv_path = save_to_excel_final(matched_df)

        get_all_file_ids_from_folder(GOOGLE_DRIVE_FOLDER_ID2)

        uploaded_excel = None
        uploaded_csv = None

        # Step 1: Upload the new files first
        file_metadata_csv = {
            'name': new_csv_name,
            'parents': [GOOGLE_DRIVE_FOLDER_ID2]
        }
        with open(temp_csv_path, 'rb') as csv_file:
            media_csv = MediaIoBaseUpload(csv_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            uploaded_csv = drive_service.files().create(
                body=file_metadata_csv, media_body=media_csv, 
                fields='id, webViewLink', supportsAllDrives=True
            ).execute()

        if upload_excel_true:
            file_metadata_excel = {
                'name': new_excel_name,
                'parents': [GOOGLE_DRIVE_FOLDER_ID2]
            }
            with open(temp_excel_path, 'rb') as excel_file:
                media_excel = MediaIoBaseUpload(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                uploaded_excel = drive_service.files().create(
                    body=file_metadata_excel, media_body=media_excel, 
                    fields='id, webViewLink', supportsAllDrives=True
                ).execute()


        # Return the new file links
        return {
            "Excel_File_Link": uploaded_excel.get("webViewLink") if uploaded_excel else "Excel upload skipped",
            "CSV_File_Link": uploaded_csv.get("webViewLink")
        }

    except Exception as e:
        logger.error("Error in creating a new file" + str(e))